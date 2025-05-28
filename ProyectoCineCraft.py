#ProyectoEDA
import openpyxl
import os
import pickle
import tkinter as tk
from tkinter import messagebox

pilaPeliculasVistas = []
listaPeliculasPendientes=[] 

#Registro
def usuarionuevo():
    def guardar():
        nombreguardar=ingresonombre.get()
        usuarioguardar=ingresarusuario.get()
        guardarcontraseña=ingresocontraseña.get()
        hoja=baseusuarios.active
        hoja.append([nombreguardar,usuarioguardar,guardarcontraseña])
        baseusuarios.save("Datos de usuarios.xlsx")
        messagebox.showinfo("Atención","Usuario registrado correctamente")
        ventanaregistro.destroy()
    ventanaregistro=tk.Tk()
    ventanaregistro.title("Registro de usuario")
    ventanaregistro.geometry("500x500")
    ventanaregistro.resizable(False,False)
    ventanaregistro.configure(bg="#000020")
    entry=tk.Entry(ventanaregistro)
    etiqueta=tk.Label(ventanaregistro,text="CREAR UNA CUENTA",
                      fg="white",
                      bg="#000020",
                      font=("Century Gothic", 30,"bold")
                    )
    Nombre=tk.Label(ventanaregistro, text="Nombre",
                    fg="white",
                    bg="#000020",
                    font=("Calibri", 14, "bold"))
    usuario=tk.Label(ventanaregistro, text="Nombre de usuario",
                     fg="white",
                     bg="#000020",
                     font=("Calibri", 14, "bold"))
    crearcontraseña=tk.Label(ventanaregistro, text="Crea una contraseña",
                             fg="white",
                             bg="#000020",
                             font=("Calibri", 14, "bold"))
    registrar=tk.Button(ventanaregistro, text="Registrar",
                        fg="white",
                        bg="#000020",
                        command=guardar,
                        font=("Century Gothic",10,"bold"))

    ingresonombre=tk.Entry(ventanaregistro,width=35)
    ingresousuario=tk.Entry(ventanaregistro,width=35)
    ingresocontraseña=tk.Entry(ventanaregistro,width=35)

    etiqueta.place(x="250",y="60",anchor="center")
    Nombre.place(x="178",y="180", anchor="center")
    usuario.place(x="135",y="230", anchor="center")
    crearcontraseña.place(x="130",y="280", anchor="center")
    registrar.place(x="250",y="350", anchor="center")
    ingresonombre.place(x="340", y="180", anchor="center")
    ingresousuario.place(x="340", y="230", anchor="center")
    ingresocontraseña.place(x="340", y="280", anchor="center")
#Ingreso
def ingresar():
    nombreusuario=nombre.get()
    contraseñausuario=contraseñaguardar.get()
    def buscar(nombreusuario,contraseñausuario):
        baseusuarios=openpyxl.load_workbook("Datos de usuarios.xlsx")
        hoja=baseusuarios.active
        for fila in hoja.iter_rows(values_only=True):
            nombre,usuario,contraseña=fila
            if usuario==nombreusuario and contraseña==contraseñausuario:
                return True
        return False
    if buscar(nombreusuario,contraseñausuario):
        ventana=tk.Tk()
        ventana.title("Inicio")
        ventana.geometry("700x700")
    else:
        messagebox.showerror("Warning","Usuario o contraseña incorrectos")
 #Opcion1
def agregarPeliculaVista():
    def guardarPelicula():
        nombre=entry_nombre.get().strip()
        duracion=entry_duracion.get().strip()
        puntuacion=entry_puntuacion.get().strip()
        año=entry_año.get().strip() 
        reseña=text_reseña.get("1.0", tk.END).strip() #"1.0" linea1, caracter 0. tk.END hasta el final del texto
        if not nombre or not duracion or not puntuacion:
            messagebox.showerror("Error", "Complete todos los campos")
            return
        #Se validan los datos
        try:
            duracionNum=float(duracion)
            puntuacionNum=float(puntuacion)
            añoNum=float(año)
        except ValueError:
            messagebox.showerror("Error", "La duración, la puntuación y el año deben ser números")
            return 
        #Se agregan a la lista
        pilaPeliculasVistas.append(nombre)
        #Se guarda en exel
        archivo_excel="Películas Vistas.xlsx"
        if os.path.exists(archivo_excel):
            libro=openpyxl.load_workbook(archivo_excel)
            hoja=libro.active
        else:
            libro=openpyxl.Workbook()
            hoja=libro.active
            hoja.append(["Nombre","Duración","Puntuación","Año"])
        hoja.append([nombre,duracionNum,puntuacionNum, añoNum])
        libro.save(archivo_excel)

        #Se guarda la reseña
        nombre_archivo_reseña=f"reseña_{nombre.replace(' ','_')}.txt"
        with open(nombre_archivo_reseña, "w", encoding="utf-8") as f:
            f.write(reseña)
        messagebox.showinfo("Éxito", f"Película'{nombre}'agregada correctamente.")
        ventanaPelicula.destroy()
        
    ventanaPelicula=tk.Toplevel()
    ventanaPelicula.title("AGREGAR PELICULA VISTA")
    ventanaPelicula.geometry("400x550")
    ventanaPelicula.configure(bg="#000020")

    tk.Label(ventanaPelicula, text="Nombre de la película:",
             fg="white",
             bg="#000020",
             font=("Century Gothic", 12)).pack(pady=(20,5))
    entry_nombre=tk.Entry(ventanaPelicula, width=40)
    entry_nombre.pack()

    tk.Label(ventanaPelicula, text="Duración (min):",
             fg="white",
             bg="#000020",
             font=("Century Gothic",12)).pack(pady=(20,5))
    entry_duracion=tk.Entry(ventanaPelicula,width=40)
    entry_duracion.pack() 
        
    tk.Label(ventanaPelicula, text="Puntuación (0-10):",
             fg="white",
             bg="#000020",
             font=("Century Gothic",12)).pack(pady=(20,5))
    entry_puntuacion=tk.Entry(ventanaPelicula,width=40)
    entry_puntuacion.pack()


    tk.Label(ventanaPelicula, text="Año:",
             fg="white",
             bg="#000020",
             font=("Century Gothic",12)).pack(pady=(20,5))
    entry_año=tk.Entry(ventanaPelicula,width=40)
    entry_año.pack()
             
    tk.Label(ventanaPelicula, text="Reseña:",
             fg="white",
             bg="#000020",
             font=("Century Gothic", 12)).pack(pady=(20,5))
    text_reseña=tk.Text(ventanaPelicula, width=40, height=8)
    text_reseña.pack()         

    boton_guardar=tk.Button(ventanaPelicula, text="Guardar",
                            fg="white",
                            bg="#000020",
                            font=("Century Gothic",12, "bold"),
                            activebackground="purple",
                            command=guardarPelicula)
    boton_guardar.pack(pady=20)
#Opcion2
def agregarPeliculaPendiente():
    def guardarPelicula():
        nombre=entry_nombre.get().strip()
        duracion=entry_duracion.get().strip()
        puntuacion=entry_puntuacion.get().strip()
        if not nombre or not duracion or not puntuacion:
            messagebox.showerror("Error", "Complete todos los campos")
            return
        #Validar datos
        try:
            duracionNum=float(duracion)
            puntuacionNum=float(puntuacion)
        except ValueError:
            messagebox.showeeoe("Error", "La duración y la puntuación deben ser números")
            return
        #Se agregan a la lista
        listaPeliculasPendientes.append(nombre)
        #Se guardan en Excel
        archivo_excel="Peliculas Pendientes.xlsx"
        if os.path.exists(archivo_excel):
            libro=openpyxl.load_workbook(archivo_excel)
            hoja=libro.active
        else:
            libro=openpyxl.Workbook()
            hoja=libro.active
            hoja.append(["Nombre","Duración","Puntuacion"])
        hoja.append([nombre, duracionNum, puntuacionNum])
        libro.save(archivo_excel)
        messagebox.showinfo("Exito", f"Pelicula pendiente'{nombre}'agregada correctamente")
        ventanaPendientes.destroy() 
                        

    ventanaPendientes=tk.Toplevel()
    ventanaPendientes.title("Agregar Película Pendiente")
    ventanaPendientes.geometry("400x400")
    ventanaPendientes.configure(bg="#000020")

    tk.Label(ventanaPendientes, text="Nombre de la película:",
                fg="white",
                bg="#000020",
                font=("Century Gothic",12)).pack(pady=(20,5))
    entry_nombre=tk.Entry(ventanaPendientes, width=40)
    entry_nombre.pack()

    tk.Label(ventanaPendientes, text="Duración de la película:",
                fg="white",
                bg="#000020",
                font=("Century Gothic",12)).pack(pady=(20,5))
    entry_duracion=tk.Entry(ventanaPendientes, width=40)
    entry_duracion.pack()

    tk.Label(ventanaPendientes, text="Puntuación de la película (1-10):",
                fg="white",
                bg="#000020",
                font=("Century Gothic",12)).pack(pady=(20,5))
    entry_puntuacion=tk.Entry(ventanaPendientes, width=40)
    entry_puntuacion.pack()

    boton_guardar=tk.Button(ventanaPendientes, text="Guardar",
                            fg="white",
                            bg="#000020",
                            font=("Century Gothic", 12,"bold"),
                            activebackground="Purple", 
                            command=guardarPelicula)
    boton_guardar.pack(pady=20) 
        
    
if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # Oculta la ventana principal si no la usas ahora
    agregarPeliculaPendiente()
    root.mainloop()
    


#Ventana principal
ventana=tk.Tk()
ventana.title("CineCraft")
ventana.geometry("700x700")
ventana.resizable(False,False)
ventana.configure(bg="#000020")
if not os.path.exists("Datos de usuarios.xlsx"):
    baseusuario=openpyxl.Workbook()
    baseusuarios.save("Datos de usuarios.xlsx")
else:
    baseusuarios=openpyxl.load_workbook("Datos de usuarios.xlsx")
marca=tk.Label(ventana, text="CINECRAFT",
                  fg="white",
                  bg="#000020",
                  font=("Century Gothic",60,"bold"))

iniciasesion=tk.Label(ventana, text="Inicia sesión",
                  fg="white",
                  bg="#000020",
                  font=("Century Gothic",30,"bold"))

usuario=tk.Label(ventana, text="Usuario",
                 fg="white",
                 bg="#000020",
                 font=("Century Gothic",12,"bold"))

nombre=tk.Entry()

contraseñaingresar=tk.Label(ventana, text="Contraseña",
                    fg="white",
                    bg="#000020",
                    font=("Century Gothic",12,"bold"))

botoningresar=tk.Button(ventana, text="Ingresar",
                        fg="white",
                        bg="#000020",
                        command=ingresar,
                        font=("Century Gothic",10,"bold"))

nuevousuario=tk.Label(ventana, text="¿Aún no tienes cuenta?",
                      fg="white",
                      bg="#000020",
                      font=("Century Gothic",8,"bold"))

registrate=tk.Button(ventana, text="Regístrate",
                     fg="orange",
                     bg="#000020",
                     borderwidth=0,
                     cursor="hand2",
                     command=usuarionuevo,
                     font=("Century Gothic",8,"bold"))

contraseñaguardar=tk.Entry(ventana, show="*")
marca.pack()
iniciasesion.place(x="350", y="150", anchor="center")
usuario.place(x="350", y="225", anchor="center")
nombre.place(x="350", y="250", anchor="center")
contraseñaingresar.place(x="350", y="300", anchor="center")
contraseñaguardar.place(x="350", y="325", anchor="center")
botoningresar.place(x="350", y="375", anchor="center")
nuevousuario.place(x="325", y="590", anchor="center")
registrate.place(x="420", y="590", anchor="center")
ventana.mainloop() 














    
    
